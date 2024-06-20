import openpyxl

def process_sheet(file_name, sheet_name, contact_field, delivery_field):
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    
    # Create a new sheet
    new_sheet_name = f"{sheet_name}_processed"
    if new_sheet_name in wb.sheetnames:
        new_sheet = wb[new_sheet_name]
    else:
        new_sheet = wb.create_sheet(new_sheet_name)
    
    # Copy headers
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(row=1, column=col).value = sheet.cell(row=1, column=col).value
    new_sheet.cell(row=1, column=sheet.max_column + 1).value = "Alternate No"
    new_sheet.cell(row=1, column=sheet.max_column + 2).value = "Direct Delivery"
    
    # Find the index of the contact and delivery fields
    contact_idx = None
    delivery_idx = None
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header == contact_field:
            contact_idx = col
        elif header == delivery_field:
            delivery_idx = col
    
    # Process each row
    for row in range(2, sheet.max_row + 1):
        contact_value = sheet.cell(row=row, column=contact_idx).value
        delivery_value = sheet.cell(row=row, column=delivery_idx).value
        
        # Split contact numbers
        if contact_value:
            contact_numbers = [num.strip() for num in contact_value.split(",")]
        else:
            contact_numbers = ["", ""]
        
        # Check for "Direct" in delivery
        if delivery_value and "Direct" in delivery_value:
            direct_delivery = "Yes"
        else:
            direct_delivery = "No"
        
        # Copy original row
        for col in range(1, sheet.max_column + 1):
            new_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value
        
        # Add new columns
        new_sheet.cell(row=row, column=sheet.max_column + 1).value = contact_numbers[1] if len(contact_numbers) > 1 else ""
        new_sheet.cell(row=row, column=sheet.max_column + 2).value = direct_delivery
    
    # Save the workbook
    wb.save(file_name)

# Example usage
file_name = 'your_excel_file.xlsx'
sheet_name = 'Sheet1'
contact_field = 'ph no'
delivery_field = 'delivery'
process_sheet(file_name, sheet_name, contact_field, delivery_field)
